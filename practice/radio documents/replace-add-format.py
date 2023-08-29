import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Define US states and political parties
us_states = [state.lower() for state in ["Alabama", "Alaska", "Arizona", "Arkansas", "California", "Colorado",
             "Connecticut", "Delaware", "Florida", "Georgia", "Hawaii", "Idaho", "Illinois",
             "Indiana", "Iowa", "Kansas", "Kentucky", "Louisiana", "Maine", "Maryland",
             "Massachusetts", "Michigan", "Minnesota", "Mississippi", "Missouri", "Montana",
             "Nebraska", "Nevada", "New Hampshire", "New Jersey", "New Mexico", "New York",
             "North Carolina", "North Dakota", "Ohio", "Oklahoma", "Oregon", "Pennsylvania",
             "Rhode Island", "South Carolina", "South Dakota", "Tennessee", "Texas", "Utah",
             "Vermont", "Virginia", "Washington", "West Virginia", "Wisconsin", "Wyoming"]]

political_parties = [party.lower() for party in ["Republican", "Democratic", "Independent"]]

# Map of short forms to their long forms
title_map = {
    "Congr ": "Congressman ",
    "Sen ": "Senator ",
    "Gov ": "Governor ",
    "Cong ": "Congressman "
}

# Load the excel file
df = pd.read_excel('input.xlsx')

# Define additional terms to find and their replacements
terms_to_replace = {
    "ll": """,
    "11": """,
    "JII:JO": ":30",
    "min.": "minutes, ",
    "sec.": "seconds, ",
	"mins": "minutes, ",
	"secs": "seconds, ",
}

# Define a function to apply to every cell in the dataframe
def add_prefix(value):
    # Make sure we are dealing with a string
    if isinstance(value, str):
        # Normalize to lower case and strip spaces for comparison
        stripped_value = value.lower().strip()
        if stripped_value in us_states:
            return 'State: ' + value
        elif stripped_value in political_parties:
            return 'Party: ' + value
        # Replace short forms with their long forms
        for short_form, long_form in title_map.items():
            if short_form in value:
                value = value.replace(short_form, long_form)
                # Add a bold heading in front of the corrected text
                value = 'Correction: ' + value
    return value

# Apply the function to the DataFrame
df = df.applymap(add_prefix)

# Write the result to a new file
df.to_excel('output.xlsx', index=False)

# Open the Excel file to format the cells
wb = load_workbook('output.xlsx')
sheet = wb.active

# Define a font style for the heading
bold_font = Font(bold=True)

# Iterate through the cells in the DataFrame and format the cells with a heading
for row in sheet.iter_rows(min_row=2, values_only=True):
    for idx, value in enumerate(row):
        if 'Correction: ' in str(value):
            # Format the cell with a bold heading
            cell = sheet.cell(row=sheet.iter_rows(min_row=2).index(row) + 2, column=idx + 1)
            cell.font = bold_font

# Save the formatted Excel file
wb.save('output_formatted.xlsx')
